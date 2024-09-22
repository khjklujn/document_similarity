from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

import session
import util


def get_topic_name(topic_id: int, config: util.Config) -> str:
    query = f"""
    select
        "name"
    from articles_topic
    where
        articles_topic.id = {topic_id}
    """
    with session.SLRV2(config) as slr_v2:
        return slr_v2.session.execute(query).fetchone()[0]


def get_prototype_articles(topic_id: int, config: util.Config) -> dict:
    query = f"""
    select
        id,
        "name"
    from articles_prototype_article
    where
        articles_prototype_article.topic_id = {topic_id}
    """
    with session.SLRV2(config) as slr_v2:
        return {
            article_id: name
            for article_id, name in slr_v2.session.execute(query).fetchall()
        }


def format_header(worksheet: Worksheet, names: tuple):
    worksheet.append(names)
    for cell in worksheet['1:1']:
        cell.font = Font(bold=True)


def get_distances(topic_id: int, config: util.Config) -> list:
    query = f"""
    select
        articles_corpus_article.id,
        articles_journal."name" journal,
        articles_year."name" "year",
        case
            when articles_corpus_article.read_status = 'On Topic' then 'Yes'
            when articles_corpus_article.read_status = 'Off Topic' then 'No'
            else ''
        end read_status,
        articles_corpus_article.percent_about,
        articles_corpus_article."name" "Article",
        1 - exp(sum(ln(1 - articles_prototype_article_corpus_article.distance))) "Relevance"
    from articles_topic
    inner join articles_prototype_article_corpus_article on
        articles_prototype_article_corpus_article.topic_id = articles_topic.id
    inner join articles_corpus_article on
        articles_corpus_article.id = articles_prototype_article_corpus_article.corpus_article_id
    inner join articles_journal on
        articles_journal.id = articles_corpus_article.journal_id
    inner join articles_year on
        articles_year.id = articles_corpus_article.year_id
    where
        articles_topic.id ={topic_id}
    group by
        articles_corpus_article.id,
        articles_topic.id,
        articles_prototype_article_corpus_article.corpus_article_id,
        articles_corpus_article."name",
        articles_journal."name",
        articles_year."name"
    order by
        "Relevance" desc
    """
    with session.SLRV2(config) as slr_v2:
        return [
            (
                record[0],
                rank + 1,
                record[1],
                record[2],
                record[3],
                record[4],
                record[5],
                record[6],
            )
            for rank, record in enumerate(slr_v2.session.execute(query).fetchall())
        ]


def get_sheet_1(workbook: Workbook, topic_id: int, topic_name: str, config: util.Config):
    print(topic_name)
    sheet = workbook.create_sheet(topic_name)

    format_header(
        sheet,
        (
            'Id',
            'Rank',
            'Journal',
            'Year',
            'On Topic',
            '% About',
            'Article',
            'Distance',
        )
    )

    [
        sheet.append(
            record
        )
        for record in get_distances(topic_id, config)
    ]

    sheet.column_dimensions['G'].width = 100

    for index, cell in enumerate(sheet['H']):
        if index > 0:
            cell.style = 'Percent'

    for index, cell in enumerate(sheet['E']):
        if index > 0:
            if cell.value == 'Yes':
                cell.font = Font(color='00AF00')
            elif cell.value == 'No':
                cell.font = Font(color='AF0000')

    for index, cell in enumerate(sheet['F']):
        if index > 0:
            cell.style = 'Percent'


def get_articles_by_prototype(prototype_article_id: int, config: util.Config) -> tuple:
    query = f"""
    select
        articles_corpus_article.id,
        articles_journal."name" journal,
        articles_year."name" "year",
        case
            when articles_corpus_article.read_status = 'On Topic' then 'Yes'
            when articles_corpus_article.read_status = 'Off Topic' then 'No'
            else ''
        end read_status,
        articles_corpus_article.percent_about,
        articles_corpus_article."name" article,
        distance,
        divergence,
        alpha,
        beta
    from articles_prototype_article
    inner join articles_prototype_article_corpus_article on
        articles_prototype_article_corpus_article.prototype_article_id = articles_prototype_article.id and
        articles_prototype_article_corpus_article.prototype_article_id = {prototype_article_id}
    inner join articles_corpus_article on
        articles_corpus_article.id = articles_prototype_article_corpus_article.corpus_article_id
    inner join articles_journal on
        articles_journal.id = articles_corpus_article.journal_id
    inner join articles_year on
        articles_year.id = articles_corpus_article.year_id
    order by
        distance desc
    """

    header = (
        'Id',
        'Rank',
        'Journal',
        'Year',
        'On Topic',
        '% About',
        'Article',
        'Distance',
        'Divergence',
        'Alpha',
        'Beta'
    )
    with session.SLRV2(config) as slr_v2:
        return (
            header,
            [
                (
                    record[0],
                    rank + 1,
                    record[1],
                    record[2],
                    record[3],
                    record[4],
                    record[5],
                    record[6],
                    record[7],
                    record[8],
                    record[9],
                )
                for rank, record in enumerate(slr_v2.session.execute(query).fetchall())
            ]
        )


def get_calculation_sheet(workbook: Workbook, topic_id: int, config: util.Config):
    query = f"""
    select
        id,
        "name"
    from articles_prototype_article
    where
        topic_id = {topic_id}
    order by
        "name"
    """

    print('Calculations')
    distances = {
        article_id: {
            'rank': rank,
            'journal': journal,
            'year': year,
            'read_status': read_status,
            'percent_about': percent_about,
            'name': name,
            'distance': distance,
        }
        for article_id, rank, journal, year, read_status, percent_about, name, distance in get_distances(topic_id, config)
    }

    with session.SLRV2(config) as slr_v2:
        header_names = [
            'Id',
            'Rank',
            'Journal',
            'Year',
            'On Topic',
            '% About',
            'Article'
        ]
        individuals = []
        for prototype_article_id, name in slr_v2.session.execute(query).fetchall():
            print('  ', name)
            individuals.append(name)
            _, rows = get_articles_by_prototype(prototype_article_id, config)
            for article_id, _, _, _, _, _, _, distance, _, _, _ in rows:
                distances[article_id][name] = distance

    sheet = workbook.create_sheet('Calculations')
    format_header(sheet, tuple(header_names + individuals))
    for article_id, item in distances.items():
        out = [
            article_id,
            item['rank'],
            item['journal'],
            item['year'],
            item['read_status'],
            item['percent_about'],
            item['name'],
        ]
        for individual_name in individuals:
            out += [item[individual_name]]
        sheet.append(out)

    sheet.column_dimensions['G'].width = 100

    for index, cell in enumerate(sheet['E']):
        if index > 0:
            if cell.value == 'Yes':
                cell.font = Font(color='00AF00')
            elif cell.value == 'No':
                cell.font = Font(color='AF0000')

    for index, cell in enumerate(sheet['F']):
        if index > 0:
            cell.style = 'Percent'

    for index, row in enumerate(sheet):
        if index > 0:
            for cell in row[7:]:
                cell.style = 'Percent'
                if cell.value >= 0.5:
                    cell.font = Font(color='00AF00')
                elif cell.value <= 0.3:
                    cell.font = Font(color='FF9F9F')
                else:
                    cell.font = Font(color='AFAFAF')


def get_individual_sheets(workbook: Workbook, topic_id: int, config: util.Config):
    query = f"""
    select
        id,
        "name"
    from articles_prototype_article
    where
        topic_id = {topic_id}
    order by
        "name"
    """

    with session.SLRV2(config) as slr_v2:
        for prototype_article_id, name in slr_v2.session.execute(query).fetchall():
            print('  ', name)
            sheet = workbook.create_sheet(name[:30])
            header_names, rows = get_articles_by_prototype(prototype_article_id, config)
            format_header(
                sheet,
                header_names,
            )
            [
                sheet.append(
                    record
                )
                for record in rows
            ]

            sheet.column_dimensions['G'].width = 100

            for index, cell in enumerate(sheet['E']):
                if index > 0:
                    if cell.value == 'Yes':
                        cell.font = Font(color='00AF00')
                    elif cell.value == 'No':
                        cell.font = Font(color='AF0000')

            for index, cell in enumerate(sheet['F']):
                if index > 0:
                    cell.style = 'Percent'

            for index, cell in enumerate(sheet['H']):
                if index > 0:
                    cell.style = 'Percent'

            for index, cell in enumerate(sheet['I']):
                if index > 0:
                    cell.style = 'Percent'

            for index, cell in enumerate(sheet['J']):
                if index > 0:
                    cell.style = 'Percent'


if __name__ == '__main__':
    config = util.Config()
    topic_id = 5
    topic_name = get_topic_name(topic_id, config)

    workbook = Workbook()

    get_sheet_1(workbook, topic_id, topic_name, config)
    get_calculation_sheet(workbook, topic_id, config)
    print(topic_name)
    get_individual_sheets(workbook, topic_id, config)

    workbook.save(f'{topic_name}.xlsx')
